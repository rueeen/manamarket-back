import re
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from .scryfall_normalizer import normalize_card_description


D = Decimal

SECTION_MAP = {
    "NM": "NM",
    "EX": "LP",
    "LP": "LP",
    "VG": "MP",
    "MP": "MP",
    "HP": "HP",
    "DMG": "DMG",
}

VALID_CONDITIONS = {"NM", "LP", "MP", "HP", "DMG"}

TOLERANCE = D("0.02")

SHIPPING_LINE_TOKENS = {
    "ups",
    "worldwide saver",
    "shipping",
    "shipment",
    "delivery",
    "freight",
    "envio",
    "envío",
    "courier",
    "dhl",
    "fedex",
    "usps",
}

FEE_LINE_TOKENS = {
    "handling",
    "customs",
    "aduana",
    "tax",
    "sales tax",
    "fee",
    "duties",
    "import duties",
}


def _is_shipping_line(description: str) -> bool:
    text = str(description or "").strip().lower()
    return any(token in text for token in SHIPPING_LINE_TOKENS)


def _is_fee_line(description: str) -> bool:
    text = str(description or "").strip().lower()
    return any(token in text for token in FEE_LINE_TOKENS)


def _is_non_product_line(description: str) -> bool:
    return _is_shipping_line(description) or _is_fee_line(description)


def _to_decimal(value):
    if value is None:
        return D("0")

    text = str(value).strip().replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)

    if not text:
        return D("0")

    try:
        return D(text)
    except InvalidOperation as exc:
        raise ValidationError(f"Valor numérico inválido: {value}") from exc


def _fmt(value):
    return f"{D(value).quantize(D('0.01'))}"


def _safe_cell(row, index, default=None):
    if index is None:
        return default

    if index < 0:
        return default

    if index >= len(row):
        return default

    return row[index]


def _get_index(headers, *names):
    for name in names:
        key = name.strip().lower()

        if key in headers:
            return headers[key]

    return None


def _get_cell_by_header(row, idx, *names, default=None):
    index = _get_index(idx, *names)
    return _safe_cell(row, index, default)


def _parse_bool(value):
    text = str(value or "").strip().lower()

    return text in {
        "1",
        "true",
        "yes",
        "y",
        "si",
        "sí",
        "foil",
    }


def _normalize_condition(value, default="NM"):
    text = str(value or default).strip().upper()

    if text in SECTION_MAP:
        return SECTION_MAP[text]

    if text in VALID_CONDITIONS:
        return text

    return default


def _parse_section(first_col):
    text = str(first_col or "").strip().upper()

    for key, value in SECTION_MAP.items():
        if text == key or text.startswith(f"{key} "):
            return value

    return None


def _extract_currency(total_value):
    text = str(total_value or "")
    match = re.search(r"\b([A-Z]{3})\b", text)

    return match.group(1) if match else None


def _is_header_row(row):
    cells = [
        str(cell or "").strip().lower()
        for cell in row[:5]
    ]

    return cells[:5] == [
        "description",
        "style",
        "qty",
        "price",
        "total",
    ]


def _apply_non_product_line_to_totals(
    *,
    description,
    amount,
    totals,
    warnings,
    row_number,
):
    """
    Evita que líneas logísticas del Excel se transformen en productos.

    Ejemplos:
    - UPS Worldwide Saver
    - Shipping
    - Handling
    - Sales Tax
    - Customs
    """
    amount = D(amount or 0)

    if _is_shipping_line(description):
        if amount > 0:
            totals["shipping_original"] += amount

        warnings.append(
            f"Fila {row_number}: '{description}' fue tratado como envío, no como producto."
        )
        return

    if _is_fee_line(description):
        normalized = str(description or "").strip().lower()

        if amount > 0:
            if "tax" in normalized:
                totals["sales_tax_original"] += amount
            elif any(
                token in normalized
                for token in {"handling", "customs", "aduana", "duties", "fee"}
            ):
                totals["shipping_original"] += amount

        warnings.append(
            f"Fila {row_number}: '{description}' fue tratado como cargo logístico/impuesto, no como producto."
        )


def _parse_normalized_workbook(wb):
    if (
        "purchase_order" not in wb.sheetnames
        or "purchase_order_items" not in wb.sheetnames
    ):
        return None

    po_sheet = wb["purchase_order"]
    fields = {}

    for row in po_sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        key = _safe_cell(row, 0)
        value = _safe_cell(row, 1)

        if key:
            fields[str(key).strip().lower()] = value

    totals = {
        "subtotal_original": _to_decimal(fields.get("subtotal_original")),
        "shipping_original": _to_decimal(fields.get("shipping_original")),
        "sales_tax_original": _to_decimal(fields.get("sales_tax_original")),
        "total_original": _to_decimal(fields.get("total_original")),
    }

    currency = str(fields.get("currency") or "").strip().upper() or "USD"

    sheet = wb["purchase_order_items"]

    header_row = next(
        sheet.iter_rows(min_row=1, max_row=1),
        None,
    )

    if not header_row:
        raise ValidationError(
            "La hoja purchase_order_items no tiene encabezados."
        )

    headers = [
        str(cell.value or "").strip().lower()
        for cell in header_row
    ]
    idx = {
        header: index
        for index, header in enumerate(headers)
        if header
    }

    required_headers = {
        "raw_description",
        "qty",
        "unit_price_original",
        "total_original",
    }

    missing = sorted(required_headers - set(idx.keys()))

    if missing:
        raise ValidationError(
            f"Faltan columnas requeridas en purchase_order_items: {', '.join(missing)}"
        )

    items = []
    errors = []
    warnings = []

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(row):
            continue

        raw = str(
            _get_cell_by_header(
                row,
                idx,
                "raw_description",
                "description",
                default="",
            )
            or ""
        ).strip()

        if _is_non_product_line(raw):
            total_value = _to_decimal(
                _get_cell_by_header(
                    row,
                    idx,
                    "total_original",
                    "line_total_original",
                    "total",
                    default=0,
                )
            )

            _apply_non_product_line_to_totals(
                description=raw,
                amount=total_value,
                totals=totals,
                warnings=warnings,
                row_number=row_number,
            )
            continue

        norm = str(
            _get_cell_by_header(
                row,
                idx,
                "normalized_name",
                "normalized_card_name",
                default="",
            )
            or ""
        ).strip()

        set_name = str(
            _get_cell_by_header(
                row,
                idx,
                "set_name",
                "set_name_detected",
                default="",
            )
            or ""
        ).strip()

        condition = _normalize_condition(
            _get_cell_by_header(
                row,
                idx,
                "condition",
                "style_condition",
                default="NM",
            )
        )

        qty = int(
            _to_decimal(
                _get_cell_by_header(
                    row,
                    idx,
                    "qty",
                    "quantity",
                    "quantity_ordered",
                    default=0,
                )
            )
        )

        unit = _to_decimal(
            _get_cell_by_header(
                row,
                idx,
                "unit_price_original",
                "price",
                default=0,
            )
        )

        total = _to_decimal(
            _get_cell_by_header(
                row,
                idx,
                "total_original",
                "line_total_original",
                "total",
                default=0,
            )
        )

        foil = _parse_bool(
            _get_cell_by_header(
                row,
                idx,
                "foil",
                "is_foil",
                default=False,
            )
        )

        language = str(
            _get_cell_by_header(
                row,
                idx,
                "language",
                default="EN",
            )
            or "EN"
        ).strip().upper() or "EN"

        if not raw and not norm:
            errors.append(
                {
                    "row": row_number,
                    "error": "Debe indicar raw_description o normalized_name.",
                }
            )
            continue

        if qty <= 0:
            errors.append(
                {
                    "row": row_number,
                    "error": "quantity_ordered debe ser > 0.",
                }
            )

        if unit < 0:
            errors.append(
                {
                    "row": row_number,
                    "error": "unit_price_original no puede ser negativo.",
                }
            )

        if total < 0:
            errors.append(
                {
                    "row": row_number,
                    "error": "total_original no puede ser negativo.",
                }
            )

        if not norm and raw:
            normalized = normalize_card_description(raw)
            norm = normalized["normalized_card_name"]
            set_name = set_name or normalized["set_name_detected"]
            foil = foil or normalized["is_foil_detected"]

        expected_total = unit * qty

        if abs(expected_total - total) > TOLERANCE:
            warnings.append(
                f"Fila {row_number}: total esperado {expected_total} distinto de total informado {total}."
            )

        items.append(
            {
                "raw_description": raw,
                "normalized_card_name": norm,
                "set_name_detected": set_name,
                "style_condition": condition,
                "quantity_ordered": qty,
                "unit_price_original": _fmt(unit),
                "line_total_original": _fmt(total),
                "is_foil_detected": foil,
                "language": language,
                "scryfall_status": "pending",
            }
        )

    return {
        "currency": currency,
        "totals": {
            key: _fmt(value)
            for key, value in totals.items()
        },
        "items": items,
        "errors": errors,
        "warnings": warnings,
    }


def parse_purchase_order_excel(file, fallback_currency="USD"):
    wb = load_workbook(file, data_only=True)

    normalized = _parse_normalized_workbook(wb)

    if normalized:
        return normalized

    sheet = wb.active

    condition = "NM"
    currency = str(fallback_currency or "USD").strip().upper() or "USD"

    items = []
    warnings = []
    errors = []

    totals = {
        "subtotal_original": D("0"),
        "shipping_original": D("0"),
        "sales_tax_original": D("0"),
        "total_original": D("0"),
    }

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(row):
            continue

        first = str(_safe_cell(row, 0, "") or "").strip()

        if not first:
            continue

        section = _parse_section(first)

        if section:
            condition = section
            continue

        if _is_header_row(row) or first.lower() == "description":
            continue

        key = first.strip().lower()

        if key in {
            "subtotal",
            "shipping",
            "sales tax",
            "total",
        }:
            raw_total = (
                _safe_cell(row, 4)
                if len(row) > 4
                else _safe_cell(row, 1, 0)
            )

            if key == "total":
                detected_currency = _extract_currency(raw_total)

                if detected_currency:
                    currency = detected_currency

            totals[f"{key.replace(' ', '_')}_original"] = _to_decimal(
                raw_total
            )
            continue

        description = first

        if _is_non_product_line(description):
            total = _to_decimal(_safe_cell(row, 4, 0))

            _apply_non_product_line_to_totals(
                description=description,
                amount=total,
                totals=totals,
                warnings=warnings,
                row_number=row_number,
            )
            continue

        qty = int(
            _to_decimal(
                _safe_cell(row, 2, 0)
            )
        )
        unit = _to_decimal(
            _safe_cell(row, 3, 0)
        )
        total = _to_decimal(
            _safe_cell(row, 4, 0)
        )
        style = _normalize_condition(
            _safe_cell(row, 1, condition),
            default=condition,
        )

        normalized = normalize_card_description(description)

        if qty <= 0:
            errors.append(
                {
                    "row": row_number,
                    "error": "quantity_ordered debe ser > 0.",
                }
            )

        if unit < 0:
            errors.append(
                {
                    "row": row_number,
                    "error": "unit_price_original no puede ser negativo.",
                }
            )

        if total < 0:
            errors.append(
                {
                    "row": row_number,
                    "error": "line_total_original no puede ser negativo.",
                }
            )

        expected_total = unit * qty

        if qty > 0 and abs(expected_total - total) > TOLERANCE:
            warnings.append(
                f"Fila {row_number}: total esperado {expected_total} distinto de total informado {total}."
            )

        items.append(
            {
                "raw_description": description,
                "normalized_card_name": normalized["normalized_card_name"],
                "set_name_detected": normalized["set_name_detected"],
                "style_condition": style,
                "quantity_ordered": qty,
                "unit_price_original": _fmt(unit),
                "line_total_original": _fmt(total),
                "is_foil_detected": normalized["is_foil_detected"],
                "language": "EN",
                "scryfall_status": "pending",
            }
        )

    calc_subtotal = sum(
        (
            D(item["line_total_original"])
            for item in items
        ),
        D("0"),
    )

    if abs(calc_subtotal - totals["subtotal_original"]) > TOLERANCE:
        warnings.append(
            f"Diferencia subtotal: items={calc_subtotal} subtotal={totals['subtotal_original']}"
        )

    calc_total = (
        totals["subtotal_original"]
        + totals["shipping_original"]
        + totals["sales_tax_original"]
    )

    if abs(calc_total - totals["total_original"]) > TOLERANCE:
        warnings.append(
            f"Diferencia total: calculado={calc_total} total={totals['total_original']}"
        )

    return {
        "currency": currency,
        "totals": {
            key: _fmt(value)
            for key, value in totals.items()
        },
        "items": items,
        "errors": errors,
        "warnings": warnings,
    }
