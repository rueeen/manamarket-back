from io import BytesIO
from unittest.mock import patch
from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import Workbook
from rest_framework.test import APIClient

from .models import KardexMovement, Product, PurchaseOrder, PurchaseOrderItem, SealedProduct, SingleCard, Supplier
from .services import parse_vendor_invoice_xlsx, resolve_scryfall_card_from_vendor


def make_xlsx(headers, rows):
    wb = Workbook(); ws = wb.active; ws.append(headers)
    for r in rows: ws.append(r)
    f = BytesIO(); wb.save(f); f.seek(0); return f

class ImportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        u = get_user_model().objects.create_user(username='admin', password='x', role='admin', is_staff=True)
        self.client.force_authenticate(u)

    @patch('products.services.get_card_by_id')
    def test_import_single_with_scryfall_id(self, mock_card):
        mock_card.return_value = {"id":"abc","name":"Lightning Bolt","set":"lea","set_name":"Alpha","collector_number":"1","prices":{"usd":"2.5"}}
        f = make_xlsx(["type","name","price_clp","scryfall_id"], [["single","Lightning Bolt",1000,"abc"]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(SingleCard.objects.count(), 1)

    def test_import_sealed_manual(self):
        f = make_xlsx(["type","name","price_clp","sealed_kind"], [["sealed","Commander Deck",20000,"precon"]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(SealedProduct.objects.count(), 1)

    def test_import_catalog_xlsx_without_file(self):
        res = self.client.post('/api/products/import-catalog-xlsx/', {}, format='multipart')
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.data['detail'], 'Debes adjuntar un archivo .xlsx')

    def test_import_catalog_xlsx_invalid_columns_returns_safe_validation_error(self):
        f = make_xlsx(["name", "price_clp"], [["Producto", 1000]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.data["detail"], "Error procesando archivo")
        self.assertIn("error", res.data)
        self.assertEqual(res.data["error"]["detail"], "Formato XLSX no reconocido")

    def test_import_catalog_xlsx_with_column_aliases(self):
        f = make_xlsx([" Tipo ", "Nombre", "Precio"], [["sealed", "Deck", 12000]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 200)
        self.assertTrue(Product.objects.filter(name='Deck', product_type='sealed').exists())

    @patch('products.services.search_cards')
    def test_reject_ambiguous_single_without_scryfall(self, mock_search):
        mock_search.return_value = [{"id":"1"},{"id":"2"}]
        f = make_xlsx(["type","name","price_clp"], [["single","Bolt",1000]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertIn('errors', res.data)

    @patch('products.services.get_card_by_id')
    def test_catalog_does_not_modify_stock(self, mock_card):
        mock_card.return_value = {"id":"abc","name":"Card","set":"lea","set_name":"Alpha","collector_number":"1","prices":{"usd":"1"}}
        p = Product.objects.create(name='Card', product_type='single', price_clp=1000, stock=7)
        f = make_xlsx(["type","name","price_clp","scryfall_id"], [["single","Card",1000,"abc"]])
        self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        p.refresh_from_db(); self.assertEqual(p.stock, 7)

    def test_import_po_increases_stock_and_kardex(self):
        p = Product.objects.create(name='Deck Box', product_type='sealed', price_clp=1000, stock=0)
        f = make_xlsx(["product_id","name","quantity","unit_cost_usd","unit_cost_clp","supplier","order_number","exchange_rate"], [[p.id,'',3,2,1500,'ABC','PO-1',900]])
        res = self.client.post('/api/purchase-orders/import-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 201)
        p.refresh_from_db(); self.assertEqual(p.stock, 0)
        self.assertFalse(KardexMovement.objects.filter(product=p, movement_type='PURCHASE_IN').exists())
        self.assertTrue(PurchaseOrderItem.objects.exists())


    @patch('products.services.search_cards')
    def test_import_single_purchase_headers_creates_single_without_price_clp(self, mock_search):
        mock_search.return_value = [{"id":"abc","name":"Lightning Bolt","set":"lea","set_name":"Alpha","collector_number":"1","prices":{"usd":"2.5"}}]
        f = make_xlsx(["name", "condition", "qty", "price_usd", "total_usd", "foil"], [["Lightning Bolt", "NM", 4, 2.5, 10, True]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 200)
        p = Product.objects.get(name='Lightning Bolt', product_type='single')
        self.assertEqual(p.price_clp, 0)
        self.assertEqual(p.stock, 0)
        self.assertEqual(SingleCard.objects.filter(product=p, condition='NM').count(), 1)

    @patch('products.services.search_cards')
    def test_import_po_with_single_purchase_headers_creates_po_items_without_stock_change(self, mock_search):
        mock_search.return_value = [{"id":"abc","name":"Counterspell","set":"2ed","set_name":"Unlimited","collector_number":"55","prices":{"usd":"1.2"}}]
        f = make_xlsx(["name", "condition", "qty", "price_usd", "total_usd", "foil"], [["Counterspell", "LP", 3, 1.2, 3.6, False]])
        res = self.client.post('/api/purchase-orders/import-xlsx/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 201)
        p = Product.objects.get(name='Counterspell', product_type='single')
        p.refresh_from_db()
        self.assertEqual(p.stock, 0)
        item = PurchaseOrderItem.objects.get(product=p)
        self.assertEqual(item.quantity_ordered, 3)

    def test_import_catalog_endpoint_accepts_post_not_405(self):
        f = make_xlsx(["type", "name", "price_clp", "sealed_kind"], [["sealed", "Bundle Box", 12000, "bundle"]])
        res = self.client.post('/api/products/import-catalog-xlsx/', {'file': f}, format='multipart')
        self.assertNotEqual(res.status_code, 405)

    def test_create_purchase_order_without_order_number_autogenerates_and_calculates_totals(self):
        supplier = Supplier.objects.create(name="Proveedor Test")
        product = Product.objects.create(name="Producto Test", product_type="sealed", price_clp=1000, stock=0)
        payload = {
            "supplier": supplier.id,
            "status": "DRAFT",
            "shipping_clp": 1000,
            "import_fees_clp": 500,
            "taxes_clp": 200,
            "order_number": "",
            "items": [
                {"product": product.id, "quantity_ordered": 2, "unit_cost_clp": 1500, "quantity_received": 0},
            ],
        }
        res = self.client.post("/api/purchase-orders/", payload, format="json")
        self.assertEqual(res.status_code, 201)
        po = PurchaseOrder.objects.get(id=res.data["id"])
        self.assertRegex(po.order_number, r"^PO-\d{8}-\d{4}$")
        self.assertEqual(po.subtotal_clp, 3000)
        self.assertEqual(po.total_clp, 4700)
        item = po.items.get(product=product)
        self.assertEqual(item.subtotal_clp, 3000)

class PurchaseOrderReceiveTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(username='worker', password='x', role='admin', is_staff=True)
        self.client.force_authenticate(self.user)

    def test_receive_purchase_order_uses_clp_when_usd_not_present(self):
        product = Product.objects.create(name='Producto CLP', product_type='sealed', price_clp=1000, stock=0)
        po = PurchaseOrder.objects.create(
            status=PurchaseOrder.Status.DRAFT,
            created_by=self.user,
            order_number='PO-CLP-1',
            subtotal_clp=4500,
        )
        item = PurchaseOrderItem.objects.create(
            purchase_order=po,
            product=product,
            quantity_ordered=3,
            quantity_received=0,
            unit_cost_usd=0,
            unit_cost_clp=1500,
            subtotal_clp=4500,
        )

        res = self.client.post(f'/api/purchase-orders/{po.id}/receive/')
        self.assertEqual(res.status_code, 200)

        item.refresh_from_db()
        product.refresh_from_db()
        po.refresh_from_db()
        self.assertEqual(item.unit_cost_clp, 1500)
        self.assertEqual(item.quantity_received, 3)
        self.assertEqual(product.stock, 3)
        self.assertEqual(po.status, PurchaseOrder.Status.RECEIVED)
        self.assertEqual(po.total_real_clp, 4500)

    def test_receive_purchase_order_allocates_additional_costs_into_unit_cost(self):
        p1 = Product.objects.create(name='Producto A', product_type='sealed', price_clp=1000, stock=0)
        p2 = Product.objects.create(name='Producto B', product_type='sealed', price_clp=1000, stock=0)
        po = PurchaseOrder.objects.create(
            status=PurchaseOrder.Status.DRAFT,
            created_by=self.user,
            order_number='PO-ALLOC-1',
            subtotal_clp=6000,
            shipping_clp=1000,
            import_fees_clp=500,
            taxes_clp=500,
        )
        i1 = PurchaseOrderItem.objects.create(
            purchase_order=po, product=p1, quantity_ordered=2, quantity_received=0, unit_cost_clp=1500, subtotal_clp=3000
        )
        i2 = PurchaseOrderItem.objects.create(
            purchase_order=po, product=p2, quantity_ordered=3, quantity_received=0, unit_cost_clp=1000, subtotal_clp=3000
        )

        res = self.client.post(f'/api/purchase-orders/{po.id}/receive/')
        self.assertEqual(res.status_code, 200)

        po.refresh_from_db()
        i1.refresh_from_db()
        i2.refresh_from_db()
        self.assertEqual(po.total_real_clp, 8000)
        self.assertEqual(i1.unit_cost_clp, 2000)
        self.assertEqual(i2.unit_cost_clp, 1333)
        self.assertEqual(i1.subtotal_clp, 4000)
        self.assertEqual(i2.subtotal_clp, 3999)

    def test_receive_purchase_order_requires_positive_cost_in_clp_or_usd(self):
        product = Product.objects.create(name='Producto Inválido', product_type='sealed', price_clp=1000, stock=0)
        po = PurchaseOrder.objects.create(status=PurchaseOrder.Status.DRAFT, created_by=self.user, order_number='PO-INVALID-1')
        PurchaseOrderItem.objects.create(
            purchase_order=po,
            product=product,
            quantity_ordered=2,
            quantity_received=0,
            unit_cost_usd=0,
            unit_cost_clp=0,
            subtotal_clp=0,
        )

        res = self.client.post(f'/api/purchase-orders/{po.id}/receive/')
        self.assertEqual(res.status_code, 400)
        self.assertIn('Costo unitario inválido', str(res.data))

    def test_create_purchase_order_auto_calculates_vat_when_empty(self):
        supplier = Supplier.objects.create(name="VAT Supplier")
        product = Product.objects.create(name="Producto VAT", product_type="sealed", price_clp=1000, stock=0)
        payload = {
            "supplier": supplier.id,
            "status": "DRAFT",
            "shipping_clp": 40000,
            "import_fees_clp": 50000,
            "taxes_clp": 0,
            "items": [{"product": product.id, "quantity_ordered": 1, "unit_cost_clp": 100000, "quantity_received": 0}],
        }
        res = self.client.post("/api/purchase-orders/", payload, format="json")
        self.assertEqual(res.status_code, 201)
        po = PurchaseOrder.objects.get(id=res.data["id"])
        self.assertEqual(po.taxes_clp, 36100)
        self.assertEqual(po.total_clp, 226100)

    def test_create_purchase_order_manual_vat_override_is_respected(self):
        supplier = Supplier.objects.create(name="VAT Override")
        product = Product.objects.create(name="Producto VAT 2", product_type="sealed", price_clp=1000, stock=0)
        payload = {
            "supplier": supplier.id,
            "status": "DRAFT",
            "shipping_clp": 1000,
            "import_fees_clp": 1000,
            "taxes_clp": 1234,
            "items": [{"product": product.id, "quantity_ordered": 1, "unit_cost_clp": 10000, "quantity_received": 0}],
        }
        res = self.client.post("/api/purchase-orders/", payload, format="json")
        self.assertEqual(res.status_code, 201)
        po = PurchaseOrder.objects.get(id=res.data["id"])
        self.assertEqual(po.taxes_clp, 1234)

    def test_product_margin_and_negative_margin(self):
        product = Product.objects.create(
            name="Producto Margen",
            product_type="sealed",
            price_clp=500,
            stock=0,
            last_purchase_cost_clp=1045,
        )
        self.assertEqual(product.cost_real_clp, 1045)
        self.assertEqual(product.margin_clp, -545)
        self.assertLess(product.margin_percentage, 0)

class PurchaseOrderImportParsingTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        u = get_user_model().objects.create_user(username='admin2', password='x', role='admin', is_staff=True)
        self.client.force_authenticate(u)

    @patch('products.purchase_order_import.search_scryfall_card')
    def test_import_purchase_order_preview_endpoint_parses_sections_and_totals(self, mock_search):
        mock_search.return_value = {"scryfall_id": "id1", "name": "Orcish Bowmasters", "set_name": "LTR", "image_large": "http://img"}
        f = make_xlsx(
            ["Description", "Style", "Qty", "Price", "Total"],
            [
                ["NM SINGLES", "", "", "", ""],
                ["The Lord of the Rings: Tales of Middle-earth: Orcish Bowmasters Foil", "", 2, "$10", "$20"],
                ["Subtotal", "", "", "", "$20"],
                ["Shipping", "", "", "", "$5"],
                ["Sales Tax", "", "", "", "$1"],
                ["Total", "", "", "", "$26"],
            ],
        )
        res = self.client.post('/api/purchase-orders/import/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.data['currency'], 'USD')
        self.assertEqual(len(res.data['preview']), 1)
        self.assertEqual(res.data['preview'][0]['normalized_name'], 'Orcish Bowmasters')
        self.assertTrue(res.data['preview'][0]['foil'])
        self.assertEqual(res.data['preview'][0]['condition'], 'NM')
        self.assertEqual(res.data['errors'], [])

    @patch('products.purchase_order_import.search_scryfall_card')
    def test_import_purchase_order_preview_detects_subtotal_inconsistency(self, mock_search):
        mock_search.return_value = None
        f = make_xlsx(["Description", "Style", "Qty", "Price", "Total"], [["Card: Test", "", 1, 2, 2], ["Subtotal", "", "", "", 3]])
        res = self.client.post('/api/purchase-orders/import/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, 200)
        self.assertTrue(any('Subtotal inconsistente' in e['error'] for e in res.data['errors']))


class VendorInvoiceParserUnitTests(TestCase):
    def test_parse_vendor_invoice_xlsx_extracts_items_and_totals(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["NM SINGLES", "", "", "", ""])
        ws.append(["Description", "Style", "Qty", "Price", "Total"])
        ws.append(["Lorwyn Foil: Blood Crypt", "", 2, "10.50", "21.00"])
        ws.append(["Subtotal", "", "", "", "21.00"])
        ws.append(["Shipping", "", "", "", "5.00"])
        ws.append(["Sales Tax", "", "", "", "0"])
        ws.append(["Total", "", "", "", "USD $26.00"])
        f = BytesIO()
        wb.save(f)
        f.seek(0)
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertEqual(len(parsed["items"]), 1)
        self.assertEqual(parsed["items"][0]["card_name"], "Blood Crypt")
        self.assertTrue(parsed["items"][0]["is_foil"])
        self.assertEqual(parsed["totals"]["total_usd"], 26)

    def test_parse_vendor_invoice_xlsx_warns_invalid_row(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["NM SINGLES", "", "", "", ""])
        ws.append(["Bad: Row", "", "abc", "x", "y"])
        f = BytesIO(); wb.save(f); f.seek(0)
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertEqual(len(parsed["items"]), 0)
        self.assertTrue(parsed["parse_warnings"])


class VendorInvoiceImportApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(username='vadmin', password='x', role='admin', is_staff=True)
        self.client.force_authenticate(self.user)

    @patch('products.views.resolve_scryfall_card_from_vendor')
    def test_import_vendor_invoice_creates_po(self, mock_resolver):
        from .models import MTGCard
        card = MTGCard.objects.create(scryfall_id="id-1", name="Blood Crypt")
        mock_resolver.return_value = (card, {"id": "id-1"}, [])
        wb = Workbook(); ws = wb.active
        ws.append(["NM SINGLES", "", "", "", ""]); ws.append(["X: Blood Crypt", "NM", 1, "2.5", "2.5"]); ws.append(["Total","","","","2.5"])
        f = BytesIO(); wb.save(f); f.seek(0)
        res = self.client.post('/api/purchase-orders/import-vendor-invoice/', {'file': f, 'supplier_name': 'CK'}, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(PurchaseOrder.objects.count(), 1)
        self.assertEqual(PurchaseOrderItem.objects.count(), 1)

    @patch('products.views.resolve_scryfall_card_from_vendor')
    def test_import_vendor_invoice_unresolved(self, mock_resolver):
        mock_resolver.return_value = (None, {"suggestions":[{"name":"A","set_code":"x","scryfall_id":"1"}]}, ["Ambiguo"])
        wb = Workbook(); ws = wb.active
        ws.append(["NM SINGLES", "", "", "", ""]); ws.append(["X: Unknown Card", "NM", 1, "2.5", "2.5"])
        f = BytesIO(); wb.save(f); f.seek(0)
        res = self.client.post('/api/purchase-orders/import-vendor-invoice/', {'file': f, 'supplier_name': 'CK'}, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data["items_unresolved"], 1)

class VendorInvoiceParserTests(TestCase):
    def _make_vendor_invoice(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        f = BytesIO()
        wb.save(f)
        f.seek(0)
        return f

    def test_parse_lotring_double_colon(self):
        f = self._make_vendor_invoice([
            ["NM SINGLES"],
            [None],
            ["Description", "Style", "Qty", "Price", "Total"],
            ["The Lord of the Rings: Tales of Middle-earth: Orcish Bowmasters", "NM", 1, 10.0, 10.0],
        ])
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertEqual(parsed["items"][0]["card_name"], "Orcish Bowmasters")
        self.assertEqual(parsed["items"][0]["set_hint"], "The Lord of the Rings: Tales of Middle-earth")

    def test_parse_foil_detection(self):
        f = self._make_vendor_invoice([
            ["NM SINGLES"],
            [None],
            ["Description", "Style", "Qty", "Price", "Total"],
            ["Lorwyn Eclipsed Foil: Blood Crypt", "NM", 1, 5.0, 5.0],
        ])
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertTrue(parsed["items"][0]["is_foil"])

    def test_parse_parenthetical_removal(self):
        f = self._make_vendor_invoice([
            ["NM SINGLES"],
            [None],
            ["Description", "Style", "Qty", "Price", "Total"],
            ["LotR Variants Foil: Sauron, the Dark Lord (0329 - Showcase)", "NM", 1, 15.0, 15.0],
        ])
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertEqual(parsed["items"][0]["card_name"], "Sauron, the Dark Lord")
        self.assertEqual(parsed["items"][0]["variant_hint"], "0329 - Showcase")

    def test_parse_totals(self):
        f = self._make_vendor_invoice([
            ["NM SINGLES"],
            [None],
            ["Description", "Style", "Qty", "Price", "Total"],
            ["Foundations Foil: Guttersnipe", "NM", 1, 1.0, 1.0],
            ["Subtotal", None, None, None, 245.69],
            ["Shipping", None, None, None, 42.78],
            ["UPS Worldwide Saver", None, None, None, None],
            ["Sales Tax", None, None, None, 0.00],
            ["Total", None, None, None, "USD $288.47"],
        ])
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertEqual(parsed["totals"]["subtotal_usd"], Decimal("245.69"))
        self.assertEqual(parsed["totals"]["shipping_usd"], Decimal("42.78"))
        self.assertEqual(parsed["totals"]["tax_usd"], Decimal("0.00"))
        self.assertEqual(parsed["totals"]["total_usd"], Decimal("288.47"))

    def test_parse_condition_mapping(self):
        f = self._make_vendor_invoice([
            ["NM SINGLES"],
            [None],
            ["Description", "Style", "Qty", "Price", "Total"],
            ["Set A: Card NM", "NM", 1, 1.0, 1.0],
            ["Set B: Card EX", "EX", 1, 1.0, 1.0],
            ["Set C: Card VG", "VG", 1, 1.0, 1.0],
            ["Set D: Card G", "G", 1, 1.0, 1.0],
            ["Set E: Card from section", "", 1, 1.0, 1.0],
        ])
        parsed = parse_vendor_invoice_xlsx(f)
        conditions = [item["condition"] for item in parsed["items"]]
        self.assertEqual(conditions, ["NM", "LP", "MP", "HP", "NM"])

    def test_parse_skip_invalid_row(self):
        f = self._make_vendor_invoice([
            ["NM SINGLES"],
            [None],
            ["Description", "Style", "Qty", "Price", "Total"],
            ["Set A: Valid Card", "NM", 1, 1.0, 1.0],
            ["Set A: Invalid Qty", "NM", "abc", 1.0, 1.0],
        ])
        parsed = parse_vendor_invoice_xlsx(f)
        self.assertEqual(len(parsed["items"]), 1)
        self.assertTrue(any("qty inválida" in warning for warning in parsed["parse_warnings"]))
