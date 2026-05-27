import logging
import re
import time
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import requests
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    Category,
    MTGCard,
    Product,
    PurchaseOrder,
    PurchaseOrderItem,
    SealedProduct,
    SingleCard,
    Supplier,
)
from .purchase_order_services import (
    get_active_exchange_rate,
    get_active_pricing_settings,
    recalculate_purchase_order,
)


SCRYFALL_BASE = "https://api.scryfall.com"
SCRYFALL_TIMEOUT = 10

logger = logging.getLogger(__name__)

D = Decimal


COLUMN_ALIASES = {
    "name": ["name", "nombre"],
    "type": ["type", "tipo", "product_type"],
    "price_clp": ["price_clp", "precio", "price"],
    "category": ["category", "categoria", "categoría"],
    "description": ["description", "descripcion", "descripción"],
    "image": ["image", "imagen", "image_url"],
    "notes": ["notes", "notas"],
    "is_active": ["is_active", "activo"],
    "condition": ["condition", "condicion", "condición", "style"],
    "language": ["language", "idioma"],
    "is_foil": ["is_foil", "foil"],
    "scryfall_id": ["scryfall_id"],
    "set_name": ["set_name", "set"],
    "sealed_kind": ["sealed_kind", "kind"],
    "set_code": ["set_code"],
}

CATALOG_REQUIRED_HEADERS = ["name", "type", "price_clp"]

SINGLE_PURCHASE_REQUIRED_HEADERS = [
    "name",
    "condition",
    "qty",
    "price_usd",
    "total_usd",
    "foil",
]


class ScryfallServiceError(Exception):
    pass


def _to_decimal(value, fallback=Decimal("0")):
    try:
        if value in (None, ""):
            return fallback
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return fallback


def _to_int(value, default=0):
    try:
        if value in (None, ""):
            return default
        return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        raise ValidationError(f"Valor numérico inválido: {value}")


def _to_bool(value, default=False):
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    return str(value).strip().lower() in {
        "true",
        "1",
        "yes",
        "y",
        "on",
        "si",
        "sí",
        "activo",
        "foil",
    }


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_condition(value):
    condition = str(value or Product.CardCondition.NM).strip().upper()

    condition_map = {
        "NM": Product.CardCondition.NM,
        "MINT": Product.CardCondition.NM,
        "M": Product.CardCondition.NM,
        "EX": Product.CardCondition.LP,
        "EXCELLENT": Product.CardCondition.LP,
        "LP": Product.CardCondition.LP,
        "VG": Product.CardCondition.MP,
        "VERY_GOOD": Product.CardCondition.MP,
        "MP": Product.CardCondition.MP,
        "G": Product.CardCondition.HP,
        "GOOD": Product.CardCondition.HP,
        "PLAYED": Product.CardCondition.HP,
        "HP": Product.CardCondition.HP,
        "PO": Product.CardCondition.DMG,
        "POOR": Product.CardCondition.DMG,
        "DMG": Product.CardCondition.DMG,
        "DAMAGED": Product.CardCondition.DMG,
    }

    normalized = condition_map.get(condition)

    if not normalized:
        raise ValidationError(f"Condición inválida: {value}")

    return normalized


def _resolve_catalog_headers(raw_headers):
    normalized_headers = [_normalize_header(header) for header in raw_headers]

    logger.info(
        "Headers recibidos en importación catálogo XLSX: %s",
        normalized_headers,
    )

    alias_to_canonical = {}

    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_to_canonical[_normalize_header(alias)] = canonical

    header_map = {}

    for header in normalized_headers:
        canonical = alias_to_canonical.get(header, header)

        if canonical not in header_map:
            header_map[canonical] = header

    missing = [
        column
        for column in CATALOG_REQUIRED_HEADERS
        if column not in header_map
    ]

    if missing:
        raise ValidationError(
            {
                "detail": "Columnas inválidas.",
                "expected": CATALOG_REQUIRED_HEADERS,
                "received": normalized_headers,
                "missing": missing,
            }
        )

    return normalized_headers, header_map


def _detect_xlsx_format(normalized_headers):
    header_set = set(normalized_headers)

    if set(CATALOG_REQUIRED_HEADERS).issubset(header_set):
        return "catalog"

    if set(SINGLE_PURCHASE_REQUIRED_HEADERS).issubset(header_set):
        return "single_purchase_items"

    raise ValidationError(
        {
            "detail": "Formato XLSX no reconocido.",
            "received_headers": normalized_headers,
            "valid_formats": {
                "catalog": CATALOG_REQUIRED_HEADERS,
                "single_purchase_items": SINGLE_PURCHASE_REQUIRED_HEADERS,
            },
        }
    )


def _request_json(path, params=None):
    url = f"{SCRYFALL_BASE}{path}"

    try:
        response = requests.get(
            url,
            params=params or {},
            timeout=SCRYFALL_TIMEOUT,
            headers={
                "User-Agent": "MTG-Ecommerce/1.0",
                "Accept": "application/json",
            },
        )
    except requests.RequestException as exc:
        raise ScryfallServiceError(
            "Error de red consultando Scryfall.") from exc

    if response.status_code == 404:
        raise ScryfallServiceError("Carta no encontrada en Scryfall.")

    if response.status_code == 429:
        raise ScryfallServiceError(
            "Scryfall limitó temporalmente las solicitudes.")

    if response.status_code >= 500:
        raise ScryfallServiceError("Scryfall temporalmente no disponible.")

    if response.status_code != 200:
        raise ScryfallServiceError(f"Error Scryfall ({response.status_code}).")

    try:
        return response.json()
    except ValueError as exc:
        raise ScryfallServiceError(
            "Respuesta inválida desde Scryfall.") from exc


def _image_uris(card_data):
    image_uris = card_data.get("image_uris") or {}

    if image_uris:
        return image_uris

    for face in card_data.get("card_faces") or []:
        if isinstance(face, dict) and face.get("image_uris"):
            return face["image_uris"]

    return {}


def _normalize_card_data(card_data):
    image_uris = _image_uris(card_data)
    released = card_data.get("released_at")

    return {
        "name": card_data.get("name", ""),
        "printed_name": card_data.get("printed_name", ""),
        "set_name": card_data.get("set_name", ""),
        "set_code": card_data.get("set", ""),
        "collector_number": card_data.get("collector_number", ""),
        "rarity": card_data.get("rarity", ""),
        "mana_cost": card_data.get("mana_cost", ""),
        "type_line": card_data.get("type_line", ""),
        "oracle_text": card_data.get("oracle_text", ""),
        "colors": card_data.get("colors") or [],
        "color_identity": card_data.get("color_identity") or [],
        "image_small": image_uris.get("small", ""),
        "image_normal": image_uris.get("normal", ""),
        "image_large": image_uris.get("large", ""),
        "scryfall_uri": card_data.get("scryfall_uri", ""),
        "released_at": date.fromisoformat(released) if released else None,
        "raw_data": card_data,
    }



def extract_usd_price(card_data, is_foil=False):
    prices = card_data.get("prices") or {}

    if is_foil:
        return (
            _to_decimal(prices.get("usd_foil"), None)
            or _to_decimal(prices.get("usd_etched"), None)
            or _to_decimal(prices.get("usd"), Decimal("0"))
        )

    return (
        _to_decimal(prices.get("usd"), None)
        or _to_decimal(prices.get("usd_foil"), None)
        or _to_decimal(prices.get("usd_etched"), Decimal("0"))
    )


def search_cards(query):
    try:
        payload = _request_json(
            "/cards/search",
            params={"q": query},
        )
    except ScryfallServiceError as exc:
        if "no encontrada" in str(exc).lower():
            return []
        raise

    return payload.get("data", [])


def get_card_by_id(scryfall_id):
    return _request_json(f"/cards/{scryfall_id}")


def get_scryfall_card_by_id(scryfall_id):
    try:
        return get_card_by_id(scryfall_id)
    except ScryfallServiceError as exc:
        logger.warning(
            "Scryfall lookup failed for scryfall_id=%s error=%s",
            scryfall_id,
            exc,
        )
        raise ValidationError(f"Scryfall no encontró la carta: {exc}") from exc


def import_card(scryfall_id):
    card_data = get_card_by_id(scryfall_id)

    card, _ = MTGCard.objects.update_or_create(
        scryfall_id=card_data["id"],
        defaults=_normalize_card_data(card_data),
    )

    return card, card_data


def _normalize_card_name(name):
    normalized_name = " ".join(
        str(name or "")
        .replace("\n", " ")
        .split()
    ).strip()

    if ":" in normalized_name:
        _, possible_name = normalized_name.split(":", 1)

        if possible_name.strip():
            normalized_name = possible_name.strip()

    return normalized_name


def _normalized_for_match(value):
    return " ".join(
        str(value or "")
        .strip()
        .lower()
        .split()
    )


def _pick_card_match(cards, normalized_name):
    if not cards:
        return None

    normalized_target = _normalized_for_match(normalized_name)

    for card_data in cards:
        if _normalized_for_match(card_data.get("name")) == normalized_target:
            return card_data

    return None


def resolve_scryfall_card(*, scryfall_id=None, name=None):
    if scryfall_id:
        card_data = get_card_by_id(str(scryfall_id).strip())

        card, _ = MTGCard.objects.update_or_create(
            scryfall_id=card_data["id"],
            defaults=_normalize_card_data(card_data),
        )

        return card, card_data, []

    normalized_name = _normalize_card_name(name)

    if not normalized_name:
        raise ValidationError("name es obligatorio para single.")

    attempted_queries = [
        f'!"{normalized_name}"',
        normalized_name,
    ]

    cards = []
    query_used = attempted_queries[0]

    for query in attempted_queries:
        query_used = query

        try:
            time.sleep(0.08)
            cards = search_cards(query)
        except ScryfallServiceError:
            continue

        if cards:
            break

    if not cards:
        raise ValidationError(
            {
                "name": normalized_name,
                "error": "No se pudo resolver la carta en Scryfall.",
                "query_used": query_used,
                "suggestion": "Agrega columna scryfall_id para importación exacta.",
            }
        )

    card_data = _pick_card_match(cards, normalized_name) or cards[0]

    if len(cards) > 1 and not _pick_card_match(cards, normalized_name):
        suggestions = [
            card.get("name")
            for card in cards[:5]
            if card.get("name")
        ]

        raise ValidationError(
            {
                "name": normalized_name,
                "error": "Resultado ambiguo en Scryfall.",
                "query_used": query_used,
                "suggestions": suggestions,
                "suggestion": "Agrega columna scryfall_id para importación exacta.",
            }
        )

    card, _ = MTGCard.objects.update_or_create(
        scryfall_id=card_data["id"],
        defaults=_normalize_card_data(card_data),
    )

    return card, card_data, ["single sin scryfall_id: se resolvió por nombre."]


VENDOR_CONDITION_MAP = {
    "NM": "NM",
    "MINT": "NM",
    "M": "NM",
    "EX": "LP",
    "EXCELLENT": "LP",
    "VG": "MP",
    "VERY GOOD": "MP",
    "G": "HP",
    "GOOD": "HP",
    "PLAYED": "HP",
    "PO": "DMG",
    "POOR": "DMG",
}


def parse_vendor_invoice_xlsx(excel_file):
    logger.info("Parsing vendor invoice xlsx")

    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active

    rows = [
        list(row or [])
        for row in sheet.iter_rows(values_only=True)
    ]

    section_pattern = re.compile(r"^[A-Z\s]+(?:SINGLES|SEALED)$")
    parenthetical_pattern = re.compile(r"\s*\(([^)]+)\)\s*$")

    total_keys = {
        "subtotal": "subtotal_usd",
        "shipping": "shipping_usd",
        "sales tax": "tax_usd",
        "tax": "tax_usd",
        "total": "total_usd",
    }

    items = []
    parse_warnings = []
    sections_found = []

    totals = {
        "subtotal_usd": Decimal("0"),
        "shipping_usd": Decimal("0"),
        "tax_usd": Decimal("0"),
        "total_usd": Decimal("0"),
    }

    sections = []

    for index, row in enumerate(rows):
        first_cell = str(row[0] if row else "" or "").strip()

        if first_cell and section_pattern.match(first_cell):
            sections.append(
                {
                    "name": first_cell,
                    "start_idx": index,
                    "header_idx": None,
                }
            )
            sections_found.append(first_cell)

    for index, section in enumerate(sections):
        search_end = (
            sections[index + 1]["start_idx"]
            if index + 1 < len(sections)
            else len(rows)
        )

        for row_index in range(section["start_idx"] + 1, search_end):
            first_cell = str(
                rows[row_index][0]
                if rows[row_index]
                else ""
                or ""
            ).strip().lower()

            if first_cell == "description":
                section["header_idx"] = row_index
                break

    for row_index, row in enumerate(rows, start=1):
        first_cell = str(row[0] if row else "" or "").strip()

        if not first_cell:
            continue

        key = first_cell.lower()

        if key in total_keys and len(row) > 4 and row[4] is not None:
            cleaned = re.sub(r"[^0-9.]", "", str(row[4]))

            if cleaned:
                totals[total_keys[key]] = Decimal(cleaned)

    for section_index, section in enumerate(sections):
        if section["header_idx"] is None:
            parse_warnings.append(
                f"Sección '{section['name']}' sin headers Description/Style/Qty/Price/Total."
            )
            continue

        section_end = (
            sections[section_index + 1]["start_idx"]
            if section_index + 1 < len(sections)
            else len(rows)
        )

        inferred_raw = section["name"].split()[0].strip().upper()
        inferred_condition = VENDOR_CONDITION_MAP.get(inferred_raw, "NM")

        for index in range(section["header_idx"] + 1, section_end):
            row = rows[index]
            description = str(row[0] if row else "" or "").strip()

            if not description or description.lower() == "description":
                continue

            low_description = description.lower()

            if low_description in total_keys or section_pattern.match(description):
                continue

            if ":" not in description:
                continue

            set_hint, card_name_raw = description.rsplit(":", 1)
            set_hint = set_hint.strip()
            card_name = card_name_raw.strip()

            variant_hint = ""
            variant_match = parenthetical_pattern.search(card_name)

            if variant_match:
                variant_hint = variant_match.group(1).strip()
                card_name = parenthetical_pattern.sub("", card_name).strip()

            style_raw = (
                str(row[1] if len(row) > 1 and row[1] is not None else "")
                .strip()
                .upper()
            )

            condition = (
                VENDOR_CONDITION_MAP.get(style_raw, inferred_condition)
                if style_raw
                else inferred_condition
            )

            try:
                qty = int(row[2])
            except (TypeError, ValueError):
                parse_warnings.append(
                    f"Fila {index + 1}: qty inválida '{row[2] if len(row) > 2 else None}'."
                )
                continue

            try:
                price_usd = Decimal(str(row[3]))
            except (InvalidOperation, TypeError, ValueError):
                parse_warnings.append(
                    f"Fila {index + 1}: price inválido '{row[3] if len(row) > 3 else None}'."
                )
                continue

            try:
                total_usd = (
                    Decimal(str(row[4]))
                    if len(row) > 4 and row[4] is not None
                    else qty * price_usd
                )
            except (InvalidOperation, TypeError, ValueError):
                parse_warnings.append(
                    f"Fila {index + 1}: total inválido '{row[4] if len(row) > 4 else None}', usando qty*price."
                )
                total_usd = qty * price_usd

            items.append(
                {
                    "row": index + 1,
                    "raw_description": description,
                    "card_name": card_name,
                    "set_hint": set_hint,
                    "variant_hint": variant_hint,
                    "is_foil": "foil" in set_hint.lower(),
                    "condition": condition,
                    "qty": qty,
                    "price_usd": price_usd,
                    "total_usd": total_usd,
                }
            )

    return {
        "items": items,
        "totals": totals,
        "sections_found": sections_found,
        "parse_warnings": parse_warnings,
    }


def resolve_scryfall_card_from_vendor(card_name, set_hint, is_foil):
    warnings = []

    cleaned = re.sub(
        r"\s+",
        " ",
        str(card_name or "").strip(),
    )

    variant_match = re.search(r"\(([^)]+)\)", cleaned)
    variant_hint = variant_match.group(1).strip() if variant_match else ""

    cleaned = re.sub(r"\([^)]*\)", "", cleaned).strip()

    queries = [f'!"{cleaned}"']

    set_token = (
        str(set_hint or "")
        .split()[0]
        .lower()
        if set_hint
        else ""
    )

    if set_token:
        queries.append(f'!"{cleaned}" {set_token}')

    queries.append(cleaned)

    all_cards = []

    for query in queries:
        try:
            time.sleep(0.1)
            cards = search_cards(query)

            if cards:
                all_cards = cards
                break
        except ScryfallServiceError as exc:
            warnings.append(f"Scryfall error query={query}: {exc}")
            continue

    if not all_cards:
        return None, None, warnings + [f"No se encontró carta para {cleaned}."]

    matched = [
        card
        for card in all_cards
        if _normalized_for_match(card.get("name")) == _normalized_for_match(cleaned)
    ] or all_cards

    if is_foil:
        foil_first = [
            card
            for card in matched
            if card.get("foil")
        ]

        if foil_first:
            matched = foil_first

    if variant_hint:
        for card in matched:
            if (
                variant_hint.lower() in str(card.get("collector_number", "")).lower()
                or variant_hint.lower() in str(card.get("frame_effects", "")).lower()
            ):
                matched = [card]
                break

    unique_names = {
        card.get("name")
        for card in matched
        if card.get("name")
    }

    if len(matched) > 1 and len(unique_names) > 1:
        suggestions = [
            {
                "name": card.get("name"),
                "set_code": card.get("set"),
                "scryfall_id": card.get("id"),
            }
            for card in matched[:3]
        ]

        return None, {"suggestions": suggestions}, warnings + [f"Ambiguo para {cleaned}."]

    card_data = matched[0]

    card, _ = MTGCard.objects.update_or_create(
        scryfall_id=card_data["id"],
        defaults=_normalize_card_data(card_data),
    )

    return card, card_data, warnings


def _resolve_category(value):
    if isinstance(value, Category):
        return value

    if not value:
        return None

    text = str(value).strip()

    return (
        Category.objects.filter(slug__iexact=text).first()
        or Category.objects.filter(name__iexact=text).first()
    )


def import_single_catalog_row(row_data):
    card, card_data, warnings = resolve_scryfall_card(
        scryfall_id=row_data.get("scryfall_id"),
        name=row_data.get("name"),
    )

    name = str(row_data.get("name") or card.name).strip()

    product, created = Product.objects.update_or_create(
        name=name,
        product_type=Product.ProductType.SINGLE,
        defaults={
            "category": _resolve_category(row_data.get("category")),
            "description": str(row_data.get("description") or card.type_line or ""),
            "price_clp": _to_int(row_data.get("price_clp"), 0),
            "image": str(
                row_data.get("image")
                or card.image_large
                or card.image_normal
                or card.image_small
                or ""
            ),
            "notes": str(row_data.get("notes") or ""),
            "is_active": _to_bool(row_data.get("is_active"), True),
        },
    )

    is_foil = _to_bool(row_data.get("is_foil"), False)

    SingleCard.objects.update_or_create(
        product=product,
        defaults={
            "mtg_card": card,
            "condition": _normalize_condition(row_data.get("condition")),
            "language": str(row_data.get("language") or "EN").upper(),
            "is_foil": is_foil,
            "edition": row_data.get("set_name") or card.set_name,
            "price_usd_reference": extract_usd_price(card_data, is_foil),
        },
    )

    return product, created, warnings


def import_single_purchase_catalog_row(row_data):
    card, card_data, warnings = resolve_scryfall_card(
        name=row_data.get("name"),
    )

    is_foil = _to_bool(row_data.get("foil"), False)

    product, created = Product.objects.update_or_create(
        name=str(row_data.get("name") or card.name).strip(),
        product_type=Product.ProductType.SINGLE,
        defaults={
            "price_clp": 0,
            "description": card.type_line or "",
            "image": card.image_large or card.image_normal or card.image_small or "",
            "is_active": True,
        },
    )

    SingleCard.objects.update_or_create(
        product=product,
        defaults={
            "mtg_card": card,
            "condition": _normalize_condition(row_data.get("condition")),
            "is_foil": is_foil,
            "language": "EN",
            "edition": card.set_name,
            "price_usd_reference": _to_decimal(
                row_data.get("price_usd"),
                Decimal("0"),
            )
            or extract_usd_price(card_data, is_foil),
        },
    )

    return product, created, warnings


def import_sealed_catalog_row(row_data):
    sealed_kind = str(row_data.get("sealed_kind") or "").strip().lower()

    if not sealed_kind:
        raise ValidationError("sealed_kind es obligatorio para type=sealed.")

    product, created = Product.objects.update_or_create(
        name=str(row_data.get("name") or "").strip(),
        product_type=Product.ProductType.SEALED,
        defaults={
            "category": _resolve_category(row_data.get("category")),
            "description": str(row_data.get("description") or ""),
            "price_clp": _to_int(row_data.get("price_clp"), 0),
            "image": str(row_data.get("image") or ""),
            "notes": str(row_data.get("notes") or ""),
            "is_active": _to_bool(row_data.get("is_active"), True),
        },
    )

    SealedProduct.objects.update_or_create(
        product=product,
        defaults={
            "sealed_kind": sealed_kind,
            "set_code": str(row_data.get("set_code") or ""),
        },
    )

    return product, created, []


def import_catalog_row(row_data):
    row_type = str(row_data.get("type") or "").strip().lower()

    if not row_type:
        raise ValidationError("type es obligatorio.")

    if not str(row_data.get("name") or "").strip():
        raise ValidationError("name es obligatorio.")

    price = _to_int(row_data.get("price_clp"), 0)

    if price < 0:
        raise ValidationError("price_clp debe ser entero >= 0.")

    if row_type == Product.ProductType.SINGLE:
        return import_single_catalog_row(row_data)

    if row_type == Product.ProductType.SEALED:
        return import_sealed_catalog_row(row_data)

    if row_type == Product.ProductType.BUNDLE:
        product, created = Product.objects.update_or_create(
            name=str(row_data.get("name")).strip(),
            product_type=Product.ProductType.BUNDLE,
            defaults={
                "category": _resolve_category(row_data.get("category")),
                "description": str(row_data.get("description") or ""),
                "price_clp": price,
                "image": str(row_data.get("image") or ""),
                "notes": str(row_data.get("notes") or ""),
                "is_active": _to_bool(row_data.get("is_active"), True),
            },
        )

        return product, created, []

    raise ValidationError("type inválido. Usa single, sealed o bundle.")


def import_catalog_from_xlsx(excel_file):
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook["catalog"] if "catalog" in workbook.sheetnames else workbook.active

    raw_headers = [
        cell.value
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]

    normalized_headers, header_map = _resolve_catalog_headers(raw_headers)

    summary = {
        "created": 0,
        "updated": 0,
        "errors": [],
        "warnings": [],
        "preview": [],
        "detected_format": "catalog",
    }

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(row):
            continue

        source_row_data = dict(zip(normalized_headers, row))

        row_data = {
            canonical: source_row_data.get(source_header)
            for canonical, source_header in header_map.items()
        }

        try:
            logger.info("Procesando fila catálogo %s", row_number)

            product, created, warnings = import_catalog_row(row_data)

            summary["created" if created else "updated"] += 1

            summary["warnings"].extend(
                [
                    {
                        "row": row_number,
                        "warning": warning,
                    }
                    for warning in warnings
                ]
            )

            summary["preview"].append(
                {
                    "row": row_number,
                    "product_id": product.id,
                    "status": "ok",
                }
            )

        except Exception as exc:
            logger.error("Error fila catálogo %s: %s", row_number, exc)

            if isinstance(exc, ValidationError) and hasattr(exc, "message_dict"):
                error_payload = {
                    "row": row_number,
                    **exc.message_dict,
                }
            else:
                error_payload = {
                    "row": row_number,
                    "error": str(exc),
                }

            summary["errors"].append(error_payload)
            summary["preview"].append(
                {
                    "row": row_number,
                    "status": "error",
                }
            )

    return summary


def _get_xlsx_sheet(workbook, preferred_sheet=None):
    if preferred_sheet and preferred_sheet in workbook.sheetnames:
        return workbook[preferred_sheet]

    return workbook.active


def _sheet_headers(sheet):
    first_row = next(
        sheet.iter_rows(min_row=1, max_row=1),
        None,
    )

    if not first_row:
        raise ValidationError("El archivo XLSX está vacío.")

    headers = [
        str(cell.value or "").strip().lower().replace(" ", "_")
        for cell in first_row
    ]

    if not any(headers):
        raise ValidationError("El archivo XLSX no tiene encabezados válidos.")

    return headers


def import_purchase_order_from_xlsx(*, excel_file, user, purchase_order_id=None):
    workbook = load_workbook(excel_file, data_only=True)
    sheet = _get_xlsx_sheet(workbook, preferred_sheet="purchase_orders")
    headers = _sheet_headers(sheet)

    detected_format = _detect_xlsx_format(headers)

    if detected_format == "catalog":
        raise ValidationError(
            "Este archivo corresponde a catálogo. Usa /api/products/import-catalog-xlsx/."
        )

    first_row = next(
        sheet.iter_rows(min_row=2, max_row=2, values_only=True),
        None,
    )

    if not first_row:
        raise ValidationError("El XLSX no contiene filas de detalle.")

    first = dict(zip(headers, first_row))

    supplier_name = str(
        first.get("supplier")
        or "Proveedor XLSX Singles"
    ).strip()

    supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

    currency = str(first.get("currency") or "USD").upper()
    exchange_rate = (
        Decimal("1")
        if currency == "CLP"
        else get_active_exchange_rate()
    )

    with transaction.atomic():
        purchase_order = (
            PurchaseOrder.objects.filter(pk=purchase_order_id).first()
            if purchase_order_id
            else None
        )

        if not purchase_order:
            order_number = str(
                first.get("order_number")
                or f"XLSX-{timezone.now().strftime('%Y%m%d%H%M%S')}"
            )

            purchase_order = PurchaseOrder.objects.create(
                supplier=supplier,
                order_number=order_number,
                created_by=user,
                status=PurchaseOrder.Status.DRAFT,
                original_currency=currency,
                exchange_rate_snapshot_clp=exchange_rate,
                source_store=str(first.get("source_store") or "XLSX"),
            )

        summary = {
            "rows_processed": 0,
            "errors": [],
            "preview": [],
        }

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(row):
                continue

            row_data = dict(zip(headers, row))
            summary["rows_processed"] += 1

            try:
                qty = _to_int(row_data.get("qty"), 0)

                if qty <= 0:
                    raise ValidationError("qty debe ser entero > 0.")

                product = None

                if row_data.get("product_id"):
                    product = Product.objects.filter(
                        pk=row_data.get("product_id")).first()

                if not product and row_data.get("name"):
                    product = Product.objects.filter(
                        name=str(row_data.get("name")).strip()
                    ).first()

                if not product:
                    product, _, _ = import_single_purchase_catalog_row(
                        row_data)

                if not product:
                    raise ValidationError(
                        "No se pudo resolver product_id/name.")

                unit_price_original = _to_decimal(
                    row_data.get("price_usd"),
                    Decimal("0"),
                )

                line_total_original = _to_decimal(
                    row_data.get("total_usd"),
                    unit_price_original * qty,
                )

                PurchaseOrderItem.objects.create(
                    purchase_order=purchase_order,
                    product=product,
                    raw_description=str(row_data.get("name") or product.name),
                    normalized_card_name=str(
                        row_data.get("name") or product.name),
                    style_condition=_normalize_condition(
                        row_data.get("condition")),
                    quantity_ordered=qty,
                    quantity_received=0,
                    unit_price_original=unit_price_original,
                    line_total_original=line_total_original,
                )

                summary["preview"].append(
                    {
                        "row": row_number,
                        "status": "ok",
                        "product_id": product.id,
                    }
                )

            except Exception as exc:
                summary["errors"].append(
                    {
                        "row": row_number,
                        "error": str(exc),
                    }
                )
                summary["preview"].append(
                    {
                        "row": row_number,
                        "status": "error",
                    }
                )

        subtotal_original = sum(
            (
                item.line_total_original
                for item in purchase_order.items.all()
            ),
            Decimal("0"),
        )

        purchase_order.subtotal_original = subtotal_original
        purchase_order.total_original = subtotal_original
        purchase_order.save(
            update_fields=[
                "subtotal_original",
                "total_original",
                "updated_at",
            ]
        )

        recalculate_purchase_order(purchase_order)

    return purchase_order, summary


def calculate_price_clp(usd_price, is_foil=False):
    usd = _to_decimal(usd_price)
    settings = get_active_pricing_settings()

    raw_clp = usd * settings.usd_to_clp

    return {
        "usd": float(usd),
        "is_foil": is_foil,
        "clp_sugerido": int(raw_clp),
    }


def calculate_suggested_sale_price(product, unit_cost_clp=None):
    unit_cost = int(unit_cost_clp or 0)

    return {
        "suggested_price_clp": max(unit_cost, int(product.price_clp or 0)),
        "min_price_clp": unit_cost,
        "source": "MANUAL",
    }
